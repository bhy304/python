import os
from os import listdir
import glob
from openpyxl import Workbook, load_workbook

ROOT_PATH = os.getcwd() # C:\Users\aithe_백하연\Desktop\python
FILE_PATH = os.path.dirname(os.path.abspath(__file__)) # c:\Users\aithe_백하연\Desktop\python\examples

txt_files_list = glob.glob(f'{ROOT_PATH}/txt/*')
xml_files_list = glob.glob(f'{ROOT_PATH}/xml/*')

result_xlsx = Workbook()
result_sheet = result_xlsx.active

for txt in range(len(txt_files_list)):
    with open(txt_files_list[txt], 'r', encoding='utf-8') as f:
        contents = f.readlines()
        # print('txt_name: ', txt_files_list[txt])
        # print('txt_contents: ', contents)
    f.close()

    for row in txt_xlsx.iter_rows():
        print(row[0])
#         row_data = []
#         for cell in row:
#             row_data.append(txt_files_list[txt], contents[cell])
        
#         result_sheet.append(row_data)

# result_xlsx.save(f'{FILE_PATH}/txt_result.xlsx')

# if len(txt_files_list) == len(xml_files_list):
#     for txt in range(len(txt_files_list)):
#         print(txt_files_list[txt])

#         with open('./check_txt.csv', 'w', newline='', encoding='utf-8') as f:
#             w = csv.writer(f)
#             for idx in range(len(txt_files_list)):
#                 w.writerow(txt_files_list[idx])
#         f.close()

# for txt in txt_files_list:
#     print(txt)