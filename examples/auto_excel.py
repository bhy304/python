import os
from os import listdir
from openpyxl import load_workbook, Workbook

ROOT_PATH = os.getcwd()
# print(f'ROOT_PATH : {ROOT_PATH}')
CUR_PATH = os.path.dirname(os.path.abspath(__file__))
# print(f'CUR_PATH : {CUR_PATH}')
path = f'{ROOT_PATH}/xlsx_files'
files = listdir(path)

result_xlsx = Workbook()
result_sheet = result_xlsx.active

for myfile in files:
    if myfile[-4:] != 'xlsx':
        continue

    tg_xlsx = load_workbook(os.path.join(path, myfile), read_only=True)
    tg_sheet = tg_xlsx.active

    for row in tg_sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)

        result_sheet.append(row_data)

result_xlsx.save(f'{CUR_PATH}/result.xlsx')