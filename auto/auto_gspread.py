import os
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

# 스프레드시트 인증하기
scope = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive'
]
CUR_PATH = os.path.dirname(os.path.abspath(__file__))
json_file_name = f'{CUR_PATH}/auto-python-294505-1522e84e7264.json'
credentials = ServiceAccountCredentials.from_json_keyfile_name(filename=json_file_name)
gs = gspread.authorize(credentials)
# 스프레드시트 문서 생성
spread = gs.create('check')
# share(개인계정이메일, 고유 대상의 유형, 공유 수준)
spread.share('bhy0512@gmail.com', perm_type='user', role='owner')

# 데이터 파일 리스트 가져오기
file_list = os.listdir(CUR_PATH)

for fname in file_list:
    if fname[-4:] != 'xlsx':
        continue
    file_path = os.path.join(CUR_PATH, fname)
    # print(file_path)

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    row_count = ws.max_row # 해당 시트의 행 수
    col_count = ws.max_column # 해당 시트의 열 수

    # 스프레드시트에 시트를 생성하고 셀 가져오기
    worksheet = spread.add_worksheet(fname, row_count, col_count)
    cells = worksheet.range('A1:' + get_column_letter(col_count) + str(row_count))

    # 스프레드시트에 데이터 반영
    idx = 0
    for row in ws.iter_rows():
        for cell in row:
            cells[idx].value = str(cell.value)
            idx = idx + 1

    worksheet.update_cells(cells)

# spread.share('bhy0315@gmail.com', perm_type='user', role='writer')